import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import openpyxl as op
matplotlib.rcParams['text.usetex'] = True


# U Cep B

wb = op.load_workbook(filename='./U_Cep_B.xlsx')
sheet = wb['Sheet1']

JD_UTC_B = [sheet.cell(i+1, 1).value for i in range(24)]

for i in range(24):
    JD_UTC_B[i] -= (59000 + 1/3)

Src_Sky_B = [sheet.cell(i+1, 2).value for i in range(24)]

Noise_B = [sheet.cell(i+1, 3).value for i in range(24)]

Ref_Src_Sky_B_1 = [sheet.cell(i+1, 4).value for i in range(24)]

Noise_Ref_Src_Sky_B_1 = [sheet.cell(i+1, 5).value for i in range(24)]

Ref_Src_Sky_B_2 = [sheet.cell(i+1, 6).value for i in range(24)]

Noise_Ref_Src_Sky_B_2 = [sheet.cell(i+1, 7).value for i in range(24)]

Ref_Src_Sky_B_3 = [sheet.cell(i+1, 8).value for i in range(24)]

Noise_Ref_Src_Sky_B_3 = [sheet.cell(i+1, 9).value for i in range(24)]

const_B_1 = [10.81+2.5*np.log10(i) for i in Ref_Src_Sky_B_1]

sigma_const_B_1 = [((2.5*Noise_Ref_Src_Sky_B_1[i]/Ref_Src_Sky_B_1[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_B_2 = [10.67+2.5*np.log10(i) for i in Ref_Src_Sky_B_2]

sigma_const_B_2 = [((2.5*Noise_Ref_Src_Sky_B_2[i]/Ref_Src_Sky_B_2[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_B_3 = [10.63+2.5*np.log10(i) for i in Ref_Src_Sky_B_3]

sigma_const_B_3 = [((2.5*Noise_Ref_Src_Sky_B_3[i]/Ref_Src_Sky_B_3[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_B = [(const_B_1[i]+const_B_2[i]+const_B_3[i])/3 for i in range(24)]

sigma_const_B = [(sigma_const_B_1[i]+sigma_const_B_2[i]+sigma_const_B_3[i])/3 for i in range(24)]

sigma_avg_const_B = [(((const_B_1[i]-const_B[i])**2+(const_B_2[i]-const_B[i])**2+(const_B_3[i]-const_B[i])**2)/6)**0.5 for i in range(24)]

mag_B = [-2.5*np.log10(Src_Sky_B[i])+const_B[i] for i in range(24)]

Noise_mag_B = [2.5*Noise_B[i]/Src_Sky_B[i] for i in range(24)]

Noise_B = [(sigma_avg_const_B[i]**2+Noise_mag_B[i]**2)**0.5 for i in range(24)]

fig, ax = plt.subplots(figsize=(16, 10), dpi=300)
ax.set(xlabel=r'$\mathrm{JD_{UTC}}-2459000$',
       ylabel=r'$m_{\mathrm{star}}/\,\mathrm{mag}$')
ax.set_xlim(165.98, 166.12)
ax.set_ylim(8.6, 10.2)
ax.invert_yaxis()
ax.grid(True, linestyle='-.')
ax.errorbar(JD_UTC_B, mag_B, yerr=Noise_B, fmt='o', color='g', mfc='b', mec='r', ms=2, lw=0.5, capsize=1.5, mew=0.5, label=r'$\mathrm{U\ Cep\ (B\ band)}$')
ax.legend(loc='upper right')
plt.savefig(fname='U_Cep_B.png', dpi=300)

# U Cep V

wb = op.load_workbook(filename='./U_Cep_V.xlsx')
sheet = wb['Sheet1']

JD_UTC_V = [sheet.cell(i+1, 1).value for i in range(24)]

for i in range(24):
    JD_UTC_V[i] -= (59000 + 1/3)

Src_Sky_V = [sheet.cell(i+1, 2).value for i in range(24)]

Noise_V = [sheet.cell(i+1, 3).value for i in range(24)]

Ref_Src_Sky_V_1 = [sheet.cell(i+1, 4).value for i in range(24)]

Noise_Ref_Src_Sky_V_1 = [sheet.cell(i+1, 5).value for i in range(24)]

Ref_Src_Sky_V_2 = [sheet.cell(i+1, 6).value for i in range(24)]

Noise_Ref_Src_Sky_V_2 = [sheet.cell(i+1, 7).value for i in range(24)]

Ref_Src_Sky_V_3 = [sheet.cell(i+1, 8).value for i in range(24)]

Noise_Ref_Src_Sky_V_3 = [sheet.cell(i+1, 9).value for i in range(24)]

const_V_1 = [10.56+2.5*np.log10(i) for i in Ref_Src_Sky_V_1]

sigma_const_V_1 = [((2.5*Noise_Ref_Src_Sky_V_1[i]/Ref_Src_Sky_V_1[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_V_2 = [10.47+2.5*np.log10(i) for i in Ref_Src_Sky_V_2]

sigma_const_V_2 = [((2.5*Noise_Ref_Src_Sky_V_2[i]/Ref_Src_Sky_V_2[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_V_3 = [10.15+2.5*np.log10(i) for i in Ref_Src_Sky_V_3]

sigma_const_V_3 = [((2.5*Noise_Ref_Src_Sky_V_3[i]/Ref_Src_Sky_V_3[i])**2 + 0.01**2/3)**0.5 for i in range(24)]

const_V = [(const_V_1[i]+const_V_2[i]+const_V_3[i])/3 for i in range(24)]

sigma_const_V = [(sigma_const_V_1[i]+sigma_const_V_2[i]+sigma_const_V_3[i])/3 for i in range(24)]

sigma_avg_const_V = [(((const_V_1[i]-const_V[i])**2+(const_V_2[i]-const_V[i])**2+(const_V_3[i]-const_V[i])**2)/6)**0.5 for i in range(24)]

mag_V = [-2.5*np.log10(Src_Sky_V[i])+const_V[i] for i in range(24)]

Noise_mag_V = [2.5*Noise_V[i]/Src_Sky_V[i] for i in range(24)]

Noise_V = [(sigma_avg_const_V[i]**2+Noise_mag_V[i]**2)**0.5 for i in range(24)]

fig, ax = plt.subplots(figsize=(16, 10), dpi=300)
ax.set(xlabel=r'$\mathrm{JD_{UTC}}-2459000$',
       ylabel=r'$m_{\mathrm{star}}/\,\mathrm{mag}$')
ax.set_xlim(165.98, 166.12)
ax.set_ylim(8.4, 9.4)
ax.invert_yaxis()
ax.grid(True, linestyle='-.')
ax.errorbar(JD_UTC_V, mag_V, yerr=Noise_V, fmt='o', color='g', mfc='b', mec='r', ms=2, lw=0.5, capsize=1.5, mew=0.5, label=r'$\mathrm{U\ Cep\ (V\ band)}$')
ax.legend(loc='upper right')
plt.savefig(fname='U_Cep_V.png', dpi=300)
