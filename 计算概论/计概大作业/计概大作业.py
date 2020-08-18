import vpython as vp
import numpy as np
from matplotlib import pyplot as plt
from scipy import optimize as opt
from scipy.interpolate import interp1d
import openpyxl as op


# 导入观测数据
wb = op.load_workbook(filename='H:/本地磁盘Z/PKU Courses/大一上/计算概论/作业/计概大作业/观测数据1.xlsx')
# 这个地址需要按照实际情况更改！共有两组观测数据。
s1 = wb['恒星数据']
s2 = wb['光变曲线']
s3 = wb['视向速度']

# 恒星数据
Ms = s1.cell(2, 2).value  # 恒星质量(Msun)
Rs = s1.cell(3, 2).value  # 恒星半径(Rsun)
Ts = s1.cell(4, 2).value  # 恒星表面的有效温度(K)

# 光变曲线数据
x_list = []
Mrow2 = s2.max_row
for i in range(2, Mrow2 + 1):
    c2 = s2.cell(i, 1)
    x_list.append(c2.value)  # 横坐标

y_list = []
for i in range(2, Mrow2 + 1):
    c2 = s2.cell(i, 2)
    y_list.append(c2.value)  # 纵坐标

# 视向速度数据
x_list2 = []
Mrow3 = s3.max_row
for i in range(2, Mrow3 + 1):
    c3 = s3.cell(i, 1)
    x_list2.append(c3.value)  # 横坐标

y_list2 = []
for i in range(2, Mrow3 + 1):
    c3 = s3.cell(i, 2)
    y_list2.append(c3.value)  # 纵坐标

# 光变曲线的平滑化处理
xmax1 = x_list[-1]
f = interp1d(x_list, y_list, kind='linear')
xnew = np.linspace(0, xmax1, num=len(x_list) * 100,
                   endpoint=True)  # i~i+1是0.01h
fx = list(f(xnew))  # 平滑后函数

for i in range(len(fx) - 1):
    if int(fx[i]) in range(-30, 31) and int(fx[i + 1]) not in range(-30, 31):
        t1 = i
    elif int(fx[i]) not in range(-30, 31) and int(fx[i + 1]) in range(-30, 31):
        t4 = i
tt = (t4 - t1) / 100  # 凌始外切到凌终外切的时间(h)

delta = int(min(fx))
for i in range(len(fx) - 1):
    if int(fx[i]) not in range(delta, delta + 21) and int(fx[i + 1]) in range(delta, delta + 21):
        t2 = i
    elif int(fx[i]) in range(delta, delta + 21) and int(fx[i + 1]) not in range(delta, delta + 21):
        t3 = i
tf = (t3 - t2) / 100  # 凌始内切到凌终内切的时间(h)
Δ = -1 * delta  # 归一化后的最大挡光量(ppm)


# 视向速度曲线拟合
xmax2 = x_list2[-1]
dx = xmax2/10
ymax2 = "{0:.1f}".format(
    ((max(y_list2) - min(y_list2)) * 0.5 * 100 + 30) / 100)  # y轴的上下界
ymax2 = float(ymax2)
dy2 = ymax2 / 5
x2 = np.arange(0, xmax2, 0.1)

def fmax(x, a, b, c):  # a,b,c=最大速度 v0(m/s),周期 P(day),相位 φ(2π=1)
    return a * np.sin((2 * np.pi / b) * x + c / (2 * np.pi))

fita, fitb = opt.curve_fit(fmax, x_list2, y_list2, bounds=([0, 0, 0], [ymax2, 1000, 1]))
# pylint: disable=unbalanced-tuple-unpacking
v0, P = fita[0], fita[1]


# 常数
G = 6.67384 * 10 ** (-11)
pi = 3.141592653589793
Msun = 1.989 * 10 ** 30
Rsun = 6.96 * 10 ** 8
Mearth = 5.9742 * 10 ** 24
Rearth = 6.371 * 10 ** 6
AU = 1.4959787 * 10 ** 11

# 恒星数据
Ms = Ms * Msun  # 恒星质量(Msun)
Rs = Rs * Rsun  # 恒星半径(Rsun)
Ts = Ts  # 恒星表面温度(K)
# 视向速度法的观测数据
P = P * 86400  # 行星的公转周期(day)
v0 = v0  # 视向速度的最大值(m/s)
# 凌星法的观测数据
tf = tf * 3600  # 凌始内切到凌终内切的时间(h)
tt = tt * 3600  # 凌始外切到凌终外切的时间(h)
Δ = 0.90862 * Δ * 10 ** (-6)  # 归一化后的最大挡光量(ppm),0.90862是临边昏暗的改正因子

# 计算
a = ((G * Ms * P ** 2) / (4 * pi ** 2)) ** (1 / 3)  # 行星的轨道半长轴(AU)
print('行星的轨道半长轴(AU): a=' + "{0:.2f}".format(a / AU))
print('行星的公转周期(平太阳日): P=' + "{0:.2f}".format(P / 86400))
v_p = (2 * pi * G * Ms / P) ** (1 / 3)  # 行星的公转线速度(m/s)
print('行星的公转线速度(km/s): vp=' + "{0:.2f}".format(v_p / 1000))
Rp = (Δ * Rs ** 2) ** (1 / 2)  # 行星的半径(Rearth)
print('行星的半径(Rearth): Rp=' + "{0:.2f}".format(Rp / Rearth))
k = (tt ** 2 + tf ** 2) / (tt ** 2 - tf ** 2)
b = (1 + Δ - 2 * Δ ** (1 / 2) * k) ** (1 / 2)
I = np.arccos(b * Rs / a)  # 轨道倾角(deg)
print('轨道倾角(deg): I=' + "{0:.2f}".format(I / pi * 180))
Mp = (Ms * v0) / (v_p * np.sin(I))  # 行星质量(Mearth)
print('行星质量(Mearth): Mp=' + "{0:.2f}".format(Mp / Mearth))
ρp = (3 * Mp) / (4 * pi * Rp ** 3)  # 行星密度(kg/m**3)
Tp = Ts * (Rs / (2 * a)) ** (1 / 2)  # 不考虑反照率、大气、内部热源等因素的行星热平衡表面温度(K)

print('行星密度(kg/m^3): ρp=' + "{0:.2f}".format(ρp))
if int(ρp) > 3000:
    print('此行星为岩质行星')
elif int(ρp) < 1500 or int(ρp) in range(1500, 3000) and int(Tp) >= -160:
    print('此行星为气态行星')
elif int(ρp) in range(1500, 3000) and int(Tp) < -160:
    print('此行星为冰质行星')

print('行星热平衡表面温度(K): Tp=' + "{0:.2f}".format(Tp))
if int(Tp) in range(274, 324):
    print('此行星在宜居带内')
elif int(Tp) > 324:
    print('行星表面温度过高，不在宜居带内')
else:
    print('行星表面温度过低，不在宜居带内')


# 常数
Rmercury = 0.382 * Rearth
Rvenus = 0.949 * Rearth
Rmars = 0.533 * Rearth
Rjupiter = 11.209 * Rearth
Rsaturn = 9.449 * Rearth
Ruranus = 4.007 * Rearth
Rneptune = 3.886 * Rearth

m1 = Ms  # 恒星质量
m2 = Mp  # 行星质量
R = a / 100  # 轨道半径 (缩小为原来的1/100)
b = b / 100
v_m2 = (G * m1 / R) ** 0.5  # 行星速度
r_m1 = Rs  # 恒星半径
r_m2 = Rp
r_m2_enlarge = r_m2  # 放大后行星半径
temp = Ts  # 恒星温度
period = P / 1000  # 周期
wavelength = 2897000 / temp  # 波长(nm)


# 恒星颜色
def radiation_color(wavelength):  # 以下为恒星光谱型
    if wavelength < 137.95:    # O
        return vp.vec(0.387, 0.719, 1)
    elif wavelength < 298.66:  # B
        return vp.vec(0.690, 0.883, 1)
    elif wavelength < 402.36:  # A
        return vp.vec(1, 1, 1)
    elif wavelength < 482.83:  # F
        return vp.vec(1, 0.84, 0)
    elif wavelength < 616.38:  # G
        return vp.vec(0.93, 0.703, 0.133)
    elif wavelength < 877.88:  # K
        return vp.vec(0.93, 0.461, 0)
    else:                      # M
        return vp.vec(0.801, 0.199, 0.199)

def Fg(x):  # 引力
    return -G * m1 * m2 / (x ** 2)


# 绘制第一个屏幕
scene = vp.canvas(width=600, height=400, center=vp.vec(0, 0, 0),
               background=vp.color.black, range=2 * R, caption='\n')
scene.lights = []
scene.title = '计概大作业 系外行星探测\n作者：李采真、高玮伯、张植竣\n此系统已经放大100倍,可用鼠标滚轮进一步调整缩放比例\n'\
              '第一张图为从地球视角所观察的系外行星运动 \n第二张图为将视角旋转90度后观察的运动\n'\
              '第三张图为系外行星与太阳系八大行星的大小比较\n第四张图为光变曲线\n'\
              '第五张图为径向速度曲线\n请调整下方滑块以对行星进行缩放'
# 缩放滑块

def S(s):  # 滑块显示数值的函数
    wts.text = '放大倍率 {:1.2f}'.format(s.value)

mra = int((r_m1 / r_m2) / 10) * 4  # 防止放大之后行星比恒星大
ratio = vp.slider(length=300, min=1, max=mra, bind=S)  # 滑块
wts = vp.wtext(text='放大倍率 ' + '1.000')

# 绘制第二个屏幕
scene2 = vp.canvas(width=600, height=400, center=vp.vec(0, 0, 0), \
                   background=vp.color.black, range=2 * R)
scene2.lights = []
scene2.title = '\n\n将视角旋转90度后观察的运动'
scene.select()


# 画行星绕恒星公转的动画
ball_m1 = vp.sphere(pos=vp.vector(0, 0, 0), radius=r_m1,
                 color=radiation_color(wavelength), emissive=True)  # 恒星
ball_m2 = vp.sphere(pos=vp.vector(0, -b * r_m1, R), radius=r_m2,
                 color=vp.color.orange, texture=vp.textures.stucco)  # 行星
ball_m1_v = vp.vector(-m2 * v_m2 / m1, 0, 0)
ball_m2_v = vp.vector(v_m2, 0, 0)
t = 0  # 初始时间
dt = period / 20000
stop = 1  # 滑块值(初始)
lamp1 = vp.local_light(pos=ball_m1.pos, color=radiation_color(wavelength))  # 调整发光源(恒星)位置

# 画转动视角后行星绕恒星公转的动画
scene2.select()
scene2.lights = []
trailball_1 = vp.sphere(pos=vp.vector(0, 0, 0), radius=r_m1, 
                        color=radiation_color(wavelength), emissive=True,
                     make_trail=True, retain=160)  # 画轨迹的恒星
trailball_2 = vp.sphere(pos=vp.vector(0, -R, -b * r_m1), radius=r_m2, 
                        color=vp.color.orange, texture=vp.textures.stucco,
                     make_trail=True, retain=160)  # 画轨迹的行星
trailball_1_v = vp.vector(-m2 * v_m2 / m1, 0, 0)
trailball_2_v = vp.vector(v_m2, 0, 0)
lamp2 = vp.local_light(pos=trailball_1.pos,
                    color=radiation_color(wavelength))  # 调整发光源(恒星)位置


# 比较八大行星大小
planets = {}
planets = {0: 'Mercury', 1: 'Venus', 2: 'Earth', 3: 'Mars',
           4: 'Jupiter', 5: 'Saturn', 6: 'Uranus', 7: 'Neptune'}
planet_radius = [0.382, 0.949, 1, 0.533, 11.209, 9.449, 4.007, 3.886]  # 行星半径
# 行星表面素材
planet_color = ['http://n.sinaimg.cn/sinacn/w640h450/20180303/fd67-fwnpcnt6414108.jpg',
                'http://n.sinaimg.cn/sinacn/w640h449/20180303/4faa-fwnpcnt6413801.jpg',
                vp.textures.earth,
                'http://n.sinaimg.cn/translate/300/w550h550/20191030/352f-ihqyuym2812320.jpg',
                'https://n.sinaimg.cn/translate/796/w934h662/20191123/f942-iittafr5320393.jpg',
                'https://n.sinaimg.cn/translate/762/w450h312/20191120/4f80-iipztff3174852.jpg',
                'https://n.sinaimg.cn/sinacn10122/742/w877h665/20191108/0daf-ihyxcrq8753528.jpg',
                'https://n.sinaimg.cn/translate/161/w640h321/20191123/13d0-iittafr3653787.jpg']
# 八大行星的屏幕
scene3 = vp.canvas(title='\n\n比较系外行星与八大行星大小', width=1200,
                height=800, background=vp.color.black)
scene3.select()
# 画系外行星
xx = -58 + Rp / r_m2 / 2
exoplanet = vp.sphere(radius=r_m2 / Rearth, pos=vp.vec(xx - r_m2 /
                    Rearth, 0, 0), color=vp.color.orange, texture=vp.textures.stucco)
T = vp.text(align='center', text='Exoplanet', pos=vp.vec(xx - r_m2 / Rearth, 13, 0), \
         color=vp.color.green)

# 画八大行星
for i in range(8):
    if i == 6:  # 天王星
        xx += 12
        planet = vp.sphere(radius=planet_radius[i], color=vp.vec(0.5, 1, 1), \
                           texture=vp.textures.rough)
    elif i == 5:  # 土星
        planet = vp.sphere(radius=planet_radius[i], color=vp.vec(0.8, 0.6, 0), \
                           texture=vp.textures.wood)
        rect = vp.shapes.rectangle(width=0.25 * planet_radius[i], height=0.1)
        circpath1 = vp.paths.circle(radius=1.93 * planet_radius[i])  # 土星环
        circpath2 = vp.paths.circle(radius=1.61 * planet_radius[i])  # 土星环
        ring1 = vp.extrusion(path=circpath1, shape=rect, \
                             pos=vp.vec((xx + planet_radius[i] + planet_radius[i - 1] + 4), 0, 0), \
                             color=vp.vec(1, 0.8, 0), texture=vp.textures.wood)  # 土星环
        ring2 = vp.extrusion(path=circpath2, shape=rect, \
                             pos=vp.vec((xx + planet_radius[i] + planet_radius[i - 1] + 4), 0, 0),
                             color=vp.vec(1, 0.8, 0), texture=vp.textures.wood)  # 土星环
        ring1.rotate(angle=vp.radians(20), axis=vp.vec(1, 0, 0))
        ring2.rotate(angle=vp.radians(20), axis=vp.vec(1, 0, 0))
    elif i == 7:  # 海王星
        planet = vp.sphere(radius=planet_radius[i], color=vp.vec(
            0, 0.25, 1), texture=vp.textures.rough)
    else:
        planet = vp.sphere(radius=planet_radius[i], texture=planet_color[i])
    planet.pos = vp.vec((xx + planet_radius[i] + planet_radius[i - 1] + 4), 0, 0)
    T = vp.text(align='center', text=planets[i], \
        pos=vp.vec(xx + planet_radius[i] + planet_radius[i - 1] + 4, 13, 0),
        color=vp.color.green)  # 显示行星名称
    xx = (xx + 3 + planet_radius[i] + planet_radius[i - 1])
    if i == 4:
        xx += 12


# 绘制光变曲线
g1 = vp.graph(width=1200, height=400, title='\n\n<b>Light Curve</b>',
           xtitle='Relative flux(ppm)', ytitle='phase(hours)')
f1 = vp.gcurve(color=vp.color.orange, width=2)
f2 = vp.gdots(color=vp.color.blue, radius=4)
for i in range(len(x_list)):
    f2.plot(x_list[i], y_list[i])
for i in range(len(xnew)):
    f1.plot(xnew[i], f(xnew[i]))

# 绘制视向速度曲线
g2 = vp.graph(width=1200, height=400, title='\n<b>Radial Velocity</b>',
           xtitle='Time(day)', ytitle='V<sub>r</sub>(m/s)')
f1 = vp.gcurve(color=vp.color.red, width=2)
f2 = vp.gdots(color=vp.color.blue, radius=4)
for i in range(len(x_list2)):
    f2.plot(x_list2[i], y_list2[i])
for i in range(len(x2)):
    f1.plot(x2[i], fmax(x2[i], fita[0], fita[1], fita[2]))


# 行星恒星运动
duration = period * 3 + dt * 1500  # 转三圈，考虑前0.5秒空转
while t <= duration:
    vp.rate(3000)
    if t <= (dt * 1500):  # 前0.5秒先空转
        t += dt
    else:
        if ratio.value != stop:  # 缩放行星
            # 缩放屏幕一的行星
            scene.select()
            position = [ball_m1.pos, ball_m2.pos,
                        trailball_1.pos, trailball_2.pos]
            velocity = [ball_m1_v, ball_m2_v, trailball_1_v, trailball_2_v]
            ball_m2.visible = False
            ball_m1.visible = False
            del ball_m2  # 删除行星
            del ball_m1  # 删除恒星
            r_m2_enlarge = r_m2 * ratio.value
            ball_m2 = vp.sphere(
                pos=position[1], radius=r_m2_enlarge, color=vp.color.orange, \
                    texture=vp.textures.stucco)
            ball_m1 = vp.sphere(pos=position[0], radius=r_m1, color=radiation_color(
                wavelength), emissive=True)
            ball_m1_v = velocity[0]
            ball_m2_v = velocity[1]

            # 缩放屏幕二的行星
            scene2.select()
            trailball_2.visible = False
            trailball_1.visible = False
            del trailball_1  # 删除行星
            del trailball_2  # 删除恒星
            trailball_1 = vp.sphere(pos=position[2], radius=r_m1, \
                                    color=radiation_color(wavelength), emissive=True,
                                 make_trail=True,)  # 画轨迹的恒星
            trailball_2 = vp.sphere(pos=position[3], radius=r_m2_enlarge, \
                                    color=vp.color.orange, texture=vp.textures.stucco,
                                 make_trail=True)  # 画轨迹的行星
            trailball_2_v = velocity[3]
            trailball_1_v = velocity[2]
            stop = ratio.value

        dist = ((ball_m1.pos.x - ball_m2.pos.x) ** 2 + (ball_m1.pos.y - ball_m2.pos.y) ** 2 + (
                ball_m1.pos.z - ball_m2.pos.z) ** 2) ** 0.5  # 距离
        radiavector = (ball_m2.pos - ball_m1.pos) / dist  # 距离单位向量
        Fg_vector = Fg(dist) * radiavector  # 行星所受万有引力
        ball_m2_v += Fg_vector / m2 * dt  # 行星速度改变
        ball_m2.pos = ball_m2.pos + ball_m2_v * dt  # 行星位移改变
        ball_m1_v += -Fg_vector / m1 * dt  # 恒星也受力
        ball_m1.pos = ball_m1.pos + ball_m1_v * dt  # 位移改变
        dist_trail = ((trailball_1.pos.x - trailball_2.pos.x) ** 2 + (trailball_1.pos.y - trailball_2.pos.y) ** 2 + (
            trailball_1.pos.z - trailball_2.pos.z) ** 2) ** 0.5  # 距离
        radiavector_trail = (trailball_2.pos - trailball_1.pos) / dist_trail  # 距离单位向量
        Fg_trail = Fg(dist_trail) * radiavector_trail  # 轨迹行星所受萬有引力
        trailball_2_v += Fg_trail / m2 * dt
        trailball_2.pos += trailball_2_v * dt
        trailball_1_v += -Fg_trail / m1 * dt
        trailball_1.pos += trailball_1_v * dt
        t += dt


# 用matplotlib绘制光变曲线
fig = plt.figure(figsize=(12, 7.5))
plt.plot(x_list, y_list, '+', xnew, f(xnew), '-')
plt.title('Light Curve', fontsize=24)
plt.xlabel("Phase(hours)", fontsize=16)
plt.ylabel("Relative flux(ppm)", fontsize=16)
plt.xlim((0, xmax1))
xaxis = np.arange(0, xmax1 + 2, 2)
plt.xticks(xaxis)

# 用matplotlib绘制视向速度曲线
fig = plt.figure(figsize=(12, 7.5))
plt.scatter(x_list2, y_list2, color='g', marker='+')
plt.plot(x2, fmax(x2, fita[0], fita[1], fita[2]), color='r')
plt.title('Radial Velocity', fontsize=24)
plt.xlabel("Time(day)", fontsize=16)
plt.ylabel("Vr(m/s)", fontsize=16)
plt.xlim((0, xmax2))
plt.ylim((-ymax2 - dy2, ymax2 + dy2))
xaxis = np.arange(0, xmax2 + dx, dx)
yaxis = np.arange(-ymax2 - dy2, ymax2 + 2 * dy2, dy2)
plt.xticks(xaxis)
plt.yticks(yaxis)

# 输出光变曲线和视向速度曲线
plt.show()  