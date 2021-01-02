import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import openpyxl as op
matplotlib.rcParams['text.usetex'] = True

# XZ Cyg B

wb = op.load_workbook(filename='./XZ_Cyg_B.xlsx')
sheet = wb['Sheet1']

JD_UTC_B = [sheet.cell(i+1, 1).value for i in range(30)]

C_B = 0.282395833 - (0.110520833 + 15 / 86400)
for i in range(6):
    JD_UTC_B[i] -= C_B
for i in range(6,30):
    JD_UTC_B[i] -= 1/3

phase_UTC_B = [(i-59144.7606)/0.4666-int((i-59144.7606)/0.4666) for i in JD_UTC_B]

Src_Sky_B = [sheet.cell(i+1, 2).value for i in range(30)]

Noise_B = [sheet.cell(i+1, 3).value for i in range(30)]

Ref_Src_Sky_B_1 = [sheet.cell(i+1, 4).value for i in range(30)]

Noise_Ref_Src_Sky_B_1 = [sheet.cell(i+1, 5).value for i in range(30)]

Ref_Src_Sky_B_2 = [sheet.cell(i+1, 6).value for i in range(30)]

Noise_Ref_Src_Sky_B_2 = [sheet.cell(i+1, 7).value for i in range(30)]

Ref_Src_Sky_B_3 = [sheet.cell(i+1, 8).value for i in range(30)]

Noise_Ref_Src_Sky_B_3 = [sheet.cell(i+1, 9).value for i in range(30)]

const_B_1 = [11.14+2.5*np.log10(i) for i in Ref_Src_Sky_B_1]

sigma_const_B_1 = [2.5*Noise_Ref_Src_Sky_B_1[i]/Ref_Src_Sky_B_1[i] for i in range(30)]

const_B_2 = [11.26+2.5*np.log10(i) for i in Ref_Src_Sky_B_2]

sigma_const_B_2 = [2.5*Noise_Ref_Src_Sky_B_2[i]/Ref_Src_Sky_B_2[i] for i in range(30)]

const_B_3 = [11.50+2.5*np.log10(i) for i in Ref_Src_Sky_B_3[0:6]] + [7.84+2.5*np.log10(i) for i in Ref_Src_Sky_B_3[6:21]] + [8.75+2.5*np.log10(i) for i in Ref_Src_Sky_B_3[21:30]]

sigma_const_B_3 = [2.5*Noise_Ref_Src_Sky_B_3[i]/Ref_Src_Sky_B_3[i] for i in range(30)]

const_B = [(const_B_1[i]+const_B_2[i]+const_B_3[i])/3 for i in range(30)]

sigma_const_B = [(sigma_const_B_1[i]+sigma_const_B_2[i]+sigma_const_B_3[i])/3 for i in range(30)]

sigma_avg_const_B = [(((sigma_const_B_1[i]-sigma_const_B[i])**2+(sigma_const_B_2[i]-sigma_const_B[i])**2+(sigma_const_B_3[i]-sigma_const_B[i])**2)/6)**0.5 for i in range(30)]

mag_B = [-2.5*np.log10(Src_Sky_B[i])+const_B[i] for i in range(30)]

Noise_mag_B = [2.5*Noise_B[i]/Src_Sky_B[i] for i in range(30)]

Noise_B = [(sigma_avg_const_B[i]**2+Noise_mag_B[i]**2)**0.5 for i in range(30)]

fig, ax = plt.subplots(figsize=(16, 10), dpi=300)
ax.set(xlabel=r'$\phi/\,2\pi$',
       ylabel=r'$m_{\mathrm{star}}/\,\mathrm{mag}$')
ax.set_xlim(0.0, 1.0)
ax.set_ylim(9.0, 10.8)
ax.set_yticks([9.0,9.3,9.6,9.9,10.2,10.5,10.8])
ax.invert_yaxis()
ax.grid(True, linestyle='-.')
ax.errorbar(phase_UTC_B, mag_B, yerr=Noise_B, fmt='o', color='g', mfc='b', mec='r', ms=2, lw=0.5, capsize=1.5, mew=0.5, label=r'$\mathrm{XZ\ Cyg\ (B\ band)}$')
ax.legend(loc='upper right')
plt.savefig(fname='XZ_Cyg_B.png', dpi=300)

# XZ Cyg V

wb = op.load_workbook(filename='./XZ_Cyg_V.xlsx')
sheet = wb['Sheet1']

JD_UTC_V = [sheet.cell(i+1, 1).value for i in range(30)]

C_V = 0.289652778 - (0.117777778 + 15 / 86400)
for i in range(6):
    JD_UTC_V[i] -= C_V
for i in range(6,30):
    JD_UTC_V[i] -= 1/3

phase_UTC_V = [(i-59144.7606)/0.4666-int((i-59144.7606)/0.4666) for i in JD_UTC_V]

Src_Sky_V = [sheet.cell(i+1, 2).value for i in range(30)]

Noise_V = [sheet.cell(i+1, 3).value for i in range(30)]

Ref_Src_Sky_V_1 = [sheet.cell(i+1, 4).value for i in range(30)]

Noise_Ref_Src_Sky_V_1 = [sheet.cell(i+1, 5).value for i in range(30)]

Ref_Src_Sky_V_2 = [sheet.cell(i+1, 6).value for i in range(30)]

Noise_Ref_Src_Sky_V_2 = [sheet.cell(i+1, 7).value for i in range(30)]

Ref_Src_Sky_V_3 = [sheet.cell(i+1, 8).value for i in range(30)]

Noise_Ref_Src_Sky_V_3 = [sheet.cell(i+1, 9).value for i in range(30)]

const_V_1 = [11.03+2.5*np.log10(i) for i in Ref_Src_Sky_V_1]

sigma_const_V_1 = [2.5*Noise_Ref_Src_Sky_V_1[i]/Ref_Src_Sky_V_1[i] for i in range(30)]

const_V_2 = [9.96+2.5*np.log10(i) for i in Ref_Src_Sky_V_2]

sigma_const_V_2 = [2.5*Noise_Ref_Src_Sky_V_2[i]/Ref_Src_Sky_V_2[i] for i in range(30)]

const_V_3 = [10.50+2.5*np.log10(i) for i in Ref_Src_Sky_V_3[0:21]] + [8.20+2.5*np.log10(i) for i in Ref_Src_Sky_V_3[21:30]]

sigma_const_V_3 = [2.5*Noise_Ref_Src_Sky_V_3[i]/Ref_Src_Sky_V_3[i] for i in range(30)]

const_V = [(const_V_1[i]+const_V_2[i]+const_V_3[i])/3 for i in range(30)]

sigma_const_V = [(sigma_const_V_1[i]+sigma_const_V_2[i]+sigma_const_V_3[i])/3 for i in range(30)]

sigma_avg_const_V = [(((sigma_const_V_1[i]-sigma_const_V[i])**2+(sigma_const_V_2[i]-sigma_const_V[i])**2+(sigma_const_V_3[i]-sigma_const_V[i])**2)/6)**0.5 for i in range(30)]

mag_V = [-2.5*np.log10(Src_Sky_V[i])+const_V[i] for i in range(30)]

Noise_mag_V = [2.5*Noise_V[i]/Src_Sky_V[i] for i in range(30)]

Noise_V = [(sigma_avg_const_V[i]**2+Noise_mag_V[i]**2)**0.5 for i in range(30)]

fig, ax = plt.subplots(figsize=(16, 10), dpi=300)
ax.set(xlabel=r'$\phi/\,2\pi$',
       ylabel=r'$m_{\mathrm{star}}/\,\mathrm{mag}$')
ax.set_xlim(0.0, 1.0)
ax.set_ylim(9.0, 10.2)
ax.set_yticks([9.0,9.2,9.4,9.6,9.8,10.0,10.2])
ax.invert_yaxis()
ax.grid(True, linestyle='-.')
ax.errorbar(phase_UTC_V, mag_V, yerr=Noise_V, fmt='o', color='g', mfc='b', mec='r', ms=2, lw=0.5, capsize=1.5, mew=0.5, label=r'$\mathrm{XZ\ Cyg\ (V\ band)}$')
ax.legend(loc='upper right')
plt.savefig(fname='XZ_Cyg_V.png', dpi=300)