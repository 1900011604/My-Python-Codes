# 使用python操作数据库，进行数据库的增删改查基本操作，任何数据库都可以。如果没接触过数据库，建议.csv
# 基本操作可以包括：
# 1，自行创建测试数据库、表，成绩表（学号，姓名，课程号，成绩）；课程表（课程号，课程名）；
# 2，查询课程号是“xx”的课程名；
# 3，查询课程总数；
# 4，查询姓“王”的学生个数；
# 5，查询所有平均成绩大于75的同学的学号和平均成绩；
# 6，增加课程表一条记录；
# 7，修改成绩表中，学号为“xxx”，课程号为“xxx”的成绩；
# 8，删除成绩表中，学号为“xxx”，课程号为“xxx”的成绩；


import re
import openpyxl as op
from openpyxl.styles import Alignment

wb = op.load_workbook(
    filename='./计算概论/作业/8作业1 数据库.xlsx')
s1 = wb['课程表']
s2 = wb['成绩表']

print('查询课程号是“xx”的课程名')
ClassID2 = input()
Mrow1 = s1.max_row
Mcol1 = s1.max_column
for i in range(2, Mrow1+1):
    if str(s1.cell(i, 1).value) == ClassID2:
        print(ClassID2, s1.cell(i, 2).value, end='\n\n')
print('查询课程总数')
print("课程总数为："+str(Mrow1-1), end='\n\n')

print('查询姓“王”的学生个数')
num = 0
for row in s2.rows:
    for cell in row:
        m = re.match('王\w+', str(cell.value))
        if m != None:
            num += 1
print("姓“王”的学生个数为"+str(num), end='\n\n')

print('查询所有平均成绩大于75的同学的学号和平均成绩')
Mrow2 = s2.max_row
Mcol2 = s2.max_column
sigma = 0
for i in range(2, Mrow2+1):  # 先行后列，从1开始
    for j in range(4, Mcol2+2, 2):
        a = s2.cell(i, j)
        sigma += int(a.value)
    average = sigma/3
    if average > 75:
        average = "{0:.2f}".format(average)
        print(s2.cell(i, 1).value, average)
    sigma = 0
print('')

print('增加课程表一条记录')
NewClassID = input()
NewClass = input()
print('新加入课程号：', NewClassID)
print('新加入课程：', NewClass, end='\n\n')
s1.insert_rows(Mrow1+1)
Mrow1 += 1
s1.cell(Mrow1, 1).value = NewClassID
s1.cell(Mrow1, 2).value = NewClass
for i in range(1, Mrow1+1):
    for j in range(1, Mcol1+1):
        s1.row_dimensions.height = 10
        s1.column_dimensions.width = 20.0
        s1.cell(i, j).alignment = Alignment(
            horizontal='center', vertical='center')
wb.save('C:/Users/19000/Desktop/PKU Courses/计算概论/作业/8作业1 数据库.xlsx')

print('修改成绩表中，学号为“xxx”，课程号为“xxx”的成绩')
StudentID7 = input()
ClassID7 = input()
Score = input()
print('学号:', StudentID7)
print('课程号:', ClassID7)
print('更改后分数:', Score, end='\n\n')
for i in range(1, Mrow2+1):  # 先行后列，从1开始
    for j in range(3, Mcol2+1, 2):
        x = s2.cell(i, 1).value
        y = s2.cell(i, j).value
        if str(x) == StudentID7 and str(y) == ClassID7:
            s2.cell(i, j+1).value = Score
            wb.save('C:/Users/19000/Desktop/PKU Courses/计算概论/作业/8作业1 数据库.xlsx')
            # 这个操作应该在Excel关闭的时候完成，如果已经打开了Excel，无法保存！

print('删除成绩表中，学号为“xxx”，课程号为“xxx”的成绩')
StudentID8 = input()
print('学号:', StudentID8)
for i in range(1, Mrow2+1):  # 先行后列，从1开始
    for j in range(3, Mcol2+1, 2):
        x = s2.cell(i, 1).value
        y = s2.cell(i, j).value
        if str(x) == StudentID8:
            s2.delete_rows(i)
            wb.save('C:/Users/19000/Desktop/PKU Courses/计算概论/作业/8作业1 数据库.xlsx')
            print('成绩已删除！')
            # 这个操作应该在Excel关闭的时候完成，如果已经打开了Excel，无法保存！
